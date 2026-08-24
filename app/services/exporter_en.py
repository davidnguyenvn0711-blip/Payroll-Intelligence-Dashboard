from __future__ import annotations

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


LABELS = {"employee_id":"Employee ID","full_name":"Full name","gross_pay":"Gross salary",
          "company":"Company","department":"Department","insurance_salary":"Insurance and taxable base salary",
          "bonus_pool":"Total bonus allocation","kpi_bonus":"Individual KPI bonus",
          "sales_bonus":"Sales achievement bonus","meal_allowance":"Meal allowance",
          "phone_allowance":"Phone allowance","travel_allowance":"Travel allowance","attendance_bonus":"Attendance bonus",
          "employee_insurance":"Employee insurance","pit":"Personal income tax","total_deductions":"Total deductions",
          "net_pay":"Net pay","employer_insurance":"Employer insurance","total_insurance":"Total insurance",
          "total_employer_cost":"Total employer cost","regular_hours":"Regular hours","overtime_hours":"Overtime hours",
          "base_hourly_rate":"Hourly rate","overtime_multiplier":"Overtime multiplier",
          "regular_pay":"Regular pay","overtime_hourly_rate":"Overtime hourly rate","overtime_pay":"Overtime pay",
          "final_payable":"Final payable"}


def payroll_excel(rows: list[dict]) -> bytes:
    columns = [c for c in LABELS if any(c in row for row in rows)]
    df = pd.DataFrame(rows).reindex(columns=columns).rename(columns=LABELS)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Payroll")
        ws = writer.book["Payroll"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="166534")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(32, max(12, max(len(str(c.value or "")) for c in col) + 2))
        non_money = {"employee_id","full_name","company","department","regular_hours","overtime_hours","overtime_multiplier"}
        money_headers = {label for key,label in LABELS.items() if key not in non_money}
        for cell in ws[1]:
            if cell.value in money_headers:
                for money_cell in ws.iter_cols(min_col=cell.column,max_col=cell.column,min_row=2):
                    for item in money_cell: item.number_format = '#,##0'
    return output.getvalue()


def payslip_pdf(row: dict, period: str) -> bytes:
    output = BytesIO()
    font = "Helvetica"
    try:
        pdfmetrics.registerFont(TTFont("DejaVu", "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"))
        font = "DejaVu"
    except Exception:
        pass
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = font
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=42, rightMargin=42, topMargin=42, bottomMargin=42)
    story = [Paragraph("PAYSLIP", styles["Title"]), Paragraph(f"Payroll period: {period}", styles["Normal"]), Spacer(1, 14)]
    data = [["Employee ID", str(row.get("employee_id", ""))], ["Full name", str(row.get("full_name", ""))]]
    for key in ["gross_pay","insurance_salary","bonus_pool","kpi_bonus","sales_bonus","meal_allowance","phone_allowance","travel_allowance","attendance_bonus","regular_hours","overtime_hours","base_hourly_rate","regular_pay","overtime_multiplier","overtime_hourly_rate","overtime_pay","final_payable","employer_insurance","employee_insurance","total_insurance","hourly_rate","ot_weekday_pay","ot_weekend_pay","ot_holiday_pay","allowances","bonuses","pit","total_deductions","net_pay"]:
        if key in row:
            label = LABELS.get(key, key.replace("_", " ").title())
            value = row[key]
            data.append([label, f"{value:,.0f}" if isinstance(value, (int,float)) else str(value)])
    table = Table(data, colWidths=[210, 260])
    table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("GRID",(0,0),(-1,-1),0.5,colors.grey),
                               ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#DCFCE7")),("PADDING",(0,0),(-1,-1),7)]))
    story += [table, Spacer(1, 18), Paragraph(f"Generated on: {date.today().strftime('%d/%m/%Y')}", styles["Normal"])]
    doc.build(story)
    return output.getvalue()
