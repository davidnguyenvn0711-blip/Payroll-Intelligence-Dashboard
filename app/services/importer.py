from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ALIASES = {
    "employee_id": ["mã nhân viên", "ma nhan vien", "employee id", "id người", "id"],
    "full_name": ["họ và tên", "ho va ten", "tên", "name"],
    "work_date": ["ngày làm việc", "ngày", "date"],
    "check_in": ["giờ vào", "vào", "check in"],
    "check_out": ["giờ ra", "ra", "check out"],
    "break_minutes": ["phút nghỉ", "thời gian nghỉ giữa ca", "break minutes"],
    "regular_hours": ["giờ thường", "giờ làm bình thường", "regular hours"],
    "ot_weekday_hours": ["tăng ca ngày thường", "ot weekday"],
    "ot_weekend_hours": ["tăng ca ngày nghỉ", "ot weekend"],
    "ot_holiday_hours": ["tăng ca ngày lễ", "ot holiday"],
    "night_hours": ["giờ ban đêm", "night hours"],
    "night_ot_hours": ["tăng ca ban đêm", "night ot"],
    "ot_approved": ["tăng ca được duyệt", "trạng thái phê duyệt tăng ca", "ot approved"]
}

ATTENDANCE_ALIASES = {
    "ID Người": "employee_id",
    "Tên": "full_name",
    "Bộ phận": "department",
    "Ngày": "work_date",
    "Tình trạng chuyên cần": "attendance_status",
    "Thời gian vào thực tế": "check_in",
    "Thời gian ra thực tế": "check_out",
    "Phút theo dõi": "tracked_minutes",
    "Số giờ theo dõi": "tracked_hours",
    "Cần kiểm tra": "needs_review",
}


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", str(value).strip().lower())
    return " ".join("".join(c for c in value if not unicodedata.combining(c)).replace("đ", "d").split())


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    lookup = {normalize_text(alias): target for target, aliases in ALIASES.items() for alias in aliases}
    renamed = {col: lookup.get(normalize_text(col), normalize_text(col).replace(" ", "_")) for col in df.columns}
    return df.rename(columns=renamed)


def read_timesheet(source, filename: str | None = None) -> pd.DataFrame:
    name = (filename or getattr(source, "name", str(source))).lower()
    if name.endswith(".csv"):
        df = pd.read_csv(source)
    elif name.endswith((".xlsx", ".xlsm")):
        df = pd.read_excel(source, sheet_name="Chấm công chi tiết" if _has_sheet(source, "Chấm công chi tiết") else 0)
    else:
        raise ValueError("Định dạng chưa hỗ trợ. Vui lòng dùng XLSX hoặc CSV; PDF máy chấm công cần chuyển bằng công cụ nhập PDF riêng.")
    df = normalize_columns(df.rename(columns=ATTENDANCE_ALIASES))
    if "work_date" in df:
        df["work_date"] = pd.to_datetime(df["work_date"], errors="coerce").dt.date
    return df


def _has_sheet(source, sheet_name: str) -> bool:
    if hasattr(source, "seek"):
        source.seek(0)
    try:
        book = pd.ExcelFile(source)
        return sheet_name in book.sheet_names
    finally:
        if hasattr(source, "seek"):
            source.seek(0)


def read_authoritative_payroll(source, sheet_name: str = "Sheet1") -> pd.DataFrame:
    """Đọc nguyên giá trị đã tính sẵn trong sheet lương chính thức.

    Hàm không tính lại công thức. Workbook phải được Excel lưu với cached values.
    """
    if hasattr(source, "seek"):
        source.seek(0)
    workbook = load_workbook(source, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Không tìm thấy sheet lương chính thức: {sheet_name}")
    sheet = workbook[sheet_name]
    header_row = None
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        if normalize_text(sheet.cell(row_number, 1).value) in {"id", "ma nhan vien"}:
            header_row = row_number
            break
    if header_row is None:
        raise ValueError("Không tìm thấy cột ID trong sheet lương chính thức")

    records = []
    for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
        employee_id = row[0]
        if employee_id is None:
            continue
        if not isinstance(employee_id, (int, float)) and not str(employee_id).strip().isdigit():
            continue
        record = {
            "employee_id": str(int(employee_id)) if isinstance(employee_id, (int, float)) else str(employee_id).strip(),
            "company": row[1],
            "full_name": row[2],
            "department": row[3],
            "gross_pay": row[4],
            "insurance_salary": row[5],
            "bonus_pool": row[6],
            "kpi_bonus": row[7],
            "sales_bonus": row[8],
            "meal_allowance": row[9],
            "phone_allowance": row[10],
            "travel_allowance": row[11],
            "attendance_bonus": row[12],
            "employer_insurance": row[13],
            "employee_insurance": row[14],
            "total_insurance": row[15],
            "source_mode": "Bảng lương đã xác nhận",
        }
        records.append(record)
    if not records:
        raise ValueError("Sheet lương chính thức không có dòng nhân viên hợp lệ")
    result = pd.DataFrame(records)
    if result.employee_id.duplicated().any():
        duplicated = result.loc[result.employee_id.duplicated(False), "employee_id"].tolist()
        raise ValueError(f"Mã nhân viên bị trùng trong bảng lương: {duplicated}")
    return result


def reconcile_employee_ids(timesheet: pd.DataFrame, payroll: pd.DataFrame) -> list[dict]:
    """Đối chiếu ID và tên; tên chỉ tạo cảnh báo, không dùng làm khóa ghép."""
    issues = []
    attendance = timesheet[["employee_id", "full_name"]].drop_duplicates().copy()
    attendance["employee_id"] = attendance.employee_id.astype(str).str.strip()
    payroll_names = payroll.set_index("employee_id")["full_name"].to_dict()
    attendance_ids = set(attendance.employee_id)
    payroll_ids = set(payroll.employee_id.astype(str))
    for employee_id in sorted(attendance_ids - payroll_ids):
        issues.append(_error("ID_ONLY_TIMESHEET", "Nghiêm trọng", None, employee_id, "ID có trong chấm công nhưng không có trong bảng lương chính thức."))
    for employee_id in sorted(payroll_ids - attendance_ids):
        issues.append(_error("ID_ONLY_PAYROLL", "Cảnh báo", None, employee_id, "ID có trong bảng lương nhưng không có chấm công tháng này."))
    for _, row in attendance[attendance.employee_id.isin(payroll_ids)].iterrows():
        if normalize_text(row.full_name) != normalize_text(payroll_names[row.employee_id]):
            issues.append(_error("NAME_MISMATCH", "Nghiêm trọng", None, row.employee_id, f"Tên theo ID không khớp: chấm công '{row.full_name}', bảng lương '{payroll_names[row.employee_id]}'."))
    return issues


def validate_timesheet(df: pd.DataFrame, employee_ids: set[str], max_daily_hours: float = 12) -> list[dict]:
    errors = []
    required = ["employee_id", "work_date"]
    for col in required:
        if col not in df:
            errors.append({"code":"MISSING_COLUMN","severity":"Nghiêm trọng","row":None,"employee_id":None,"message":f"Thiếu cột bắt buộc: {col}"})
    if errors:
        return errors
    duplicate_mask = df.duplicated(subset=[c for c in ["employee_id","work_date","check_in","check_out"] if c in df], keep=False)
    for idx, row in df.iterrows():
        emp = str(row.get("employee_id", "")).strip()
        if not emp or emp.lower() == "nan":
            errors.append(_error("MISSING_EMPLOYEE_ID", "Nghiêm trọng", idx, emp, "Thiếu mã nhân viên; bổ sung mã duy nhất."))
        elif emp not in employee_ids:
            errors.append(_error("UNKNOWN_EMPLOYEE", "Nghiêm trọng", idx, emp, "Mã nhân viên chưa có trong danh mục."))
        if duplicate_mask.loc[idx]:
            errors.append(_error("DUPLICATE", "Nghiêm trọng", idx, emp, "Dòng chấm công bị trùng."))
        if ("check_in" in df and pd.isna(row.get("check_in"))) != ("check_out" in df and pd.isna(row.get("check_out"))):
            errors.append(_error("MISSING_PUNCH", "Nghiêm trọng", idx, emp, "Thiếu giờ vào hoặc giờ ra; cần xác nhận thủ công."))
        hours = sum(float(row.get(c, 0) or 0) for c in ["regular_hours","ot_weekday_hours","ot_weekend_hours","ot_holiday_hours","night_hours"])
        if hours < 0 or hours > max_daily_hours:
            errors.append(_error("UNREASONABLE_HOURS", "Nghiêm trọng", idx, emp, f"Tổng {hours:g} giờ/ngày ngoài ngưỡng 0-{max_daily_hours:g}."))
        ot = sum(float(row.get(c, 0) or 0) for c in ["ot_weekday_hours","ot_weekend_hours","ot_holiday_hours","night_ot_hours"])
        if ot > 0 and str(row.get("ot_approved", "")).strip().lower() not in {"true","1","có","da duyet","đã duyệt"}:
            errors.append(_error("OT_NOT_APPROVED", "Nghiêm trọng", idx, emp, "Tăng ca chưa được phê duyệt và sẽ không được trả."))
    return errors


def _error(code, severity, idx, employee_id, message):
    return {"code":code,"severity":severity,"row":int(idx)+2,"employee_id":employee_id or None,"message":message}
